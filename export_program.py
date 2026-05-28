"""
export_program.py — Export client training programs from Firestore to Excel.

Usage:
  python3 export_program.py              # export all clients
  python3 export_program.py an           # export specific client by ID
  python3 export_program.py an joost     # export multiple clients

Output: outputs/{client_id}_program.xlsx
"""

import json
import sys
import urllib.request
import urllib.error
import urllib.parse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colors ──────────────────────────────────────────────────────────────────
BRAND_BLACK   = "FF1A1A1A"
BRAND_ACCENT  = "FFE8C547"   # yellow-gold
WHITE         = "FFFFFFFF"
LIGHT_GRAY    = "FFF5F5F5"
MID_GRAY      = "FFCCCCCC"
DARK_GRAY     = "FF555555"
PHASE_COLORS  = {
    "strength":    ("FF1C3050", WHITE),   # dark navy  / white
    "rehab":       ("FF2D5016", WHITE),   # dark green / white
    "warmup":      ("FF7B3F00", WHITE),   # dark brown / white
    "core":        ("FF3B0070", WHITE),   # dark purple/ white
    "accessories": ("FF004D40", WHITE),   # teal       / white
    "default":     ("FF2C2C2C", WHITE),   # near-black / white
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color=BRAND_BLACK, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Helvetica Neue")

def _border(style="thin", color=MID_GRAY):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _thin_bottom(color=MID_GRAY):
    s = Side(border_style="thin", color=color)
    n = Side(border_style=None)
    return Border(left=n, right=n, top=n, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── Firestore helpers ────────────────────────────────────────────────────────
def _parse_value(val):
    if "stringValue"  in val: return val["stringValue"]
    if "integerValue" in val: return int(val["integerValue"])
    if "doubleValue"  in val: return float(val["doubleValue"])
    if "booleanValue" in val: return val["booleanValue"]
    if "nullValue"    in val: return None
    if "mapValue"     in val:
        return {k: _parse_value(v) for k, v in val["mapValue"].get("fields", {}).items()}
    if "arrayValue"   in val:
        return [_parse_value(v) for v in val["arrayValue"].get("values", [])]
    return val

def fetch_client(client_id, access_token, project_id="fitness-app-a22c8"):
    encoded_id = urllib.parse.quote(client_id, safe="")
    url = (f"https://firestore.googleapis.com/v1/projects/{project_id}"
           f"/databases/(default)/documents/clients/{encoded_id}")
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {access_token}"})
    with urllib.request.urlopen(req) as resp:
        raw = json.loads(resp.read())
    if "fields" not in raw:
        return {}
    return {k: _parse_value(v) for k, v in raw["fields"].items()}

def fetch_session_loads(client_id, access_token, project_id="fitness-app-a22c8"):
    """Return dict: {session_key: {set_loads: {key: [kg, ...]}, exercise_loads: {key: kg}}}"""
    encoded_id = urllib.parse.quote(client_id, safe="")
    url = (f"https://firestore.googleapis.com/v1/projects/{project_id}"
           f"/databases/(default)/documents/clients/{encoded_id}/sessionLoads")
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {access_token}"})
    try:
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
    except Exception:
        return {}
    result = {}
    for doc in data.get("documents", []):
        session_id = doc["name"].split("/")[-1]
        fields = {k: _parse_value(v) for k, v in doc.get("fields", {}).items()}
        result[session_id] = {
            "setLoads":      fields.get("setLoads", {}),
            "exerciseLoads": fields.get("exerciseLoads", {}),
        }
    return result

def list_clients(access_token, project_id="fitness-app-a22c8"):
    url = (f"https://firestore.googleapis.com/v1/projects/{project_id}"
           f"/databases/(default)/documents/clients")
    req = urllib.request.Request(url, headers={"Authorization": f"Bearer {access_token}"})
    with urllib.request.urlopen(req) as resp:
        data = json.loads(resp.read())
    return [doc["name"].split("/")[-1] for doc in data.get("documents", [])]

# ── Sheet builder ─────────────────────────────────────────────────────────────
COLS = ["Exercise", "Sets / Reps", "Tempo", "Cue / Notes", "Load (kg)"]
COL_WIDTHS = [36, 16, 10, 42, 22]

def _phase_colors(phase_name: str):
    key = phase_name.lower()
    for k, v in PHASE_COLORS.items():
        if k in key:
            return v
    return PHASE_COLORS["default"]

def _load_label(client_id: str, session_id: str, phase_idx: int, ex_idx: int,
                session_loads: dict) -> str:
    """Return per-set load string e.g. '40 / 50 / 50 kg' or '' if no data."""
    key = f"{client_id}_{session_id}_{phase_idx}_{ex_idx}"
    s_data = session_loads.get(session_id, {})
    sets = s_data.get("setLoads", {}).get(key)
    if sets:
        return " / ".join(str(int(k) if k == int(k) else k) for k in sets) + " kg"
    best = s_data.get("exerciseLoads", {}).get(key)
    if best is not None:
        return f"{int(best) if best == int(best) else best} kg"
    return ""

def build_sheet(ws, day_label: str, phases: list, client: dict,
                session_key: str = "", session_loads: dict = None):
    ws.sheet_view.showGridLines = False

    # ── Column widths ────────────────────────────────────────────────────────
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── Day header ───────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 36
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value=day_label.upper())
    c.font      = _font(bold=True, size=14, color=BRAND_ACCENT)
    c.fill      = _fill(BRAND_BLACK)
    c.alignment = _align("left", "center")
    row += 1

    # ── Client info sub-header ───────────────────────────────────────────────
    ws.row_dimensions[row].height = 18
    ws.merge_cells(f"A{row}:E{row}")
    info = f"{client.get('name','')}  |  {client.get('level','')}  |  {client.get('goal','')}"
    c = ws.cell(row=row, column=1, value=info)
    c.font      = _font(size=9, color="FFAAAAAA", italic=True)
    c.fill      = _fill(BRAND_BLACK)
    c.alignment = _align("left", "center")
    row += 1

    # ── Spacer ───────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 6
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = _fill(BRAND_BLACK)
    row += 1

    for phase_idx, phase in enumerate(phases):
        phase_name = phase.get("name", "Phase")
        exercises  = phase.get("exercises", [])
        note       = phase.get("note", "")
        bg_hex, fg_hex = _phase_colors(phase_name)

        # ── Phase header ─────────────────────────────────────────────────────
        ws.row_dimensions[row].height = 26
        ws.merge_cells(f"A{row}:E{row}")
        phase_label = f"  {phase_name}"
        if note:
            phase_label += f"  ·  {note}"
        c = ws.cell(row=row, column=1, value=phase_label)
        c.font      = _font(bold=True, size=11, color=fg_hex)
        c.fill      = _fill(bg_hex)
        c.alignment = _align("left", "center")
        row += 1

        # ── Column headers ───────────────────────────────────────────────────
        ws.row_dimensions[row].height = 20
        for col_i, col_name in enumerate(COLS, start=1):
            c = ws.cell(row=row, column=col_i, value=col_name)
            c.font      = _font(bold=True, size=9, color=DARK_GRAY)
            c.fill      = _fill(LIGHT_GRAY)
            c.alignment = _align("left", "center")
            c.border    = _thin_bottom("FFAAAAAA")
        row += 1

        # ── Exercise rows ────────────────────────────────────────────────────
        for idx, ex in enumerate(exercises):
            ws.row_dimensions[row].height = 22
            fill_color = WHITE if idx % 2 == 0 else "FFFAFAFA"
            load_str = _load_label(
                client.get("id", ""), session_key, phase_idx, idx,
                session_loads or {}
            )
            values = [
                ex.get("name", ""),
                ex.get("setsReps", ""),
                ex.get("tempo", ""),
                ex.get("cue", ""),
                load_str,
            ]
            for col_i, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=col_i, value=val)
                c.font      = _font(size=10)
                c.fill      = _fill(fill_color)
                c.alignment = _align("left", "center", wrap=(col_i == 4))
                c.border    = _thin_bottom()
            row += 1

        # ── Gap between phases ───────────────────────────────────────────────
        ws.row_dimensions[row].height = 8
        row += 1

    # ── Footer ───────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 14
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="Pulse Fitness  ·  pulsefit.vercel.app")
    c.font      = _font(size=8, color="FFBBBBBB", italic=True)
    c.fill      = _fill(WHITE)
    c.alignment = _align("center", "center")

# ── Main export ───────────────────────────────────────────────────────────────
def export_client(client_id: str, access_token: str):
    print(f"  Fetching {client_id}...", end=" ", flush=True)
    client = fetch_client(client_id, access_token)
    client["id"] = client_id
    program = client.get("program", {})

    if not program:
        print("no program data — skipped.")
        return

    session_loads = fetch_session_loads(client_id, access_token)

    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    day_order = ["SessionA", "SessionB", "SessionC",
                 "Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    days = sorted(program.keys(),
                  key=lambda d: day_order.index(d) if d in day_order else 99)

    for day_key in days:
        day_data   = program[day_key]
        day_label  = day_data.get("label", day_key)
        phases     = day_data.get("phases", [])
        sheet_name = day_key[:31]   # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=sheet_name)
        build_sheet(ws, day_label, phases, client,
                    session_key=day_key, session_loads=session_loads)

    out_path = Path("outputs") / f"{client_id}_program.xlsx"
    wb.save(out_path)
    print(f"saved → {out_path}")
    return out_path

def _get_access_token() -> str:
    """Get a fresh access token via firebase-tools apiv2 (handles refresh automatically)."""
    import subprocess, shutil
    node = shutil.which("node")
    if not node:
        raise RuntimeError("node not found — cannot refresh Firebase token")

    script = r"""
const apiv2 = require('/Users/longchu/.npm-global/lib/node_modules/firebase-tools/lib/apiv2.js');
const fs    = require('fs');
const cfg   = JSON.parse(fs.readFileSync(process.env.HOME + '/.config/configstore/firebase-tools.json', 'utf8'));
apiv2.setRefreshToken(cfg.tokens.refresh_token);
const client = new apiv2.Client({ urlPrefix: 'https://oauth2.googleapis.com', auth: true });
// Trigger a lightweight authenticated call to force a token refresh, then read it back
const https = require('https');
const qs    = require('querystring');
// Use the same refresh mechanism apiv2 uses internally
apiv2.Client.prototype;
// Just read the refreshed token directly from the apiv2 internals
const apiv2Src = fs.readFileSync(
  '/Users/longchu/.npm-global/lib/node_modules/firebase-tools/lib/apiv2.js', 'utf8');
// apiv2 stores token after refresh — trigger a real request instead
const firestoreClient = new apiv2.Client({ urlPrefix: 'https://firestore.googleapis.com', auth: true });
firestoreClient.request({ method: 'GET',
  path: '/v1/projects/fitness-app-a22c8/databases/(default)/documents/clients?pageSize=1',
  responseType: 'json' })
  .then(() => {
    // Token is now refreshed in the configstore
    const cfg2 = JSON.parse(fs.readFileSync(process.env.HOME + '/.config/configstore/firebase-tools.json', 'utf8'));
    process.stdout.write(cfg2.tokens.access_token);
  })
  .catch(e => { process.stderr.write(e.message); process.exit(1); });
"""
    result = subprocess.run([node, "-e", script], capture_output=True, text=True, timeout=30)
    if result.returncode != 0:
        raise RuntimeError(f"Token refresh failed: {result.stderr.strip()}")
    return result.stdout.strip()

def main():
    print("Refreshing Firebase token...", end=" ", flush=True)
    access_token = _get_access_token()
    print("OK")

    target_ids = sys.argv[1:] if len(sys.argv) > 1 else None

    if target_ids is None:
        print("Fetching all client IDs...")
        target_ids = list_clients(access_token)

    print(f"Exporting {len(target_ids)} client(s):")
    for cid in target_ids:
        try:
            export_client(cid, access_token)
        except Exception as e:
            import traceback
            print(f"  ERROR {cid}: {e}")
            traceback.print_exc()

    print("Done.")

if __name__ == "__main__":
    main()
